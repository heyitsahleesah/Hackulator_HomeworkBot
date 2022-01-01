import discord
import openpyxl
import datetime
import random
import os
from dotenv import load_dotenv
import pandas as pd

client = discord.Client()

encouragements = [
    "No time like the present",
    "You can do it!",
    "I believe in you!",
    "It will be easy",
]

# use builtin function to find the date on your own computer
current_date = datetime.datetime.now()
# convert the date to a readable format with only relevant information
printable_date = current_date.strftime("%m-%d")

# location of the excel file
file = 'DiscordExcel.xlsx'

# Open the workbook
wb = openpyxl.load_workbook(file)
sheet = wb.active


@client.event
async def on_ready():
    # message to show that the bot is logged in (connected properly)
    print('We have logged in as {0.user}'.format(client))


# iterate through the worksheet on call and delete rows that are determined to be empty
async def delete_empty_rows(message):
    """
    Iterates through the worksheet and deletes rows that have no values.

    :return:
    """
    # # schedule cleanup
    # cleanup_time = datetime.time(00, 00, 00)
    # now = datetime.time.now()
    for row in sheet.rows():
        if all(cell.value for cell in row) is not True:
            sheet.delete_rows(row[0].row, 1)
    await message.channel.send('Empty rows deleted!')


# check the date (based on the date comparison function) and delete the rows of the worksheet where the date has passed
async def delete_past_homework(rows, message):
    """
    Iterates through the worksheet and deletes rows with dates that have passed.

    :return:
    """
    for row in sheet.rows:
        if int(sheet.cell(rows, 4).value[1:3]) > int(printable_date[0:2]):
            sheet.delete_rows(row, 1)
        elif int(sheet.cell(rows, 4).value[1:3]) == int(printable_date[0:2]) \
                and int(sheet.cell(rows, 4).value[4:6]) > int(printable_date[3:5]):
            sheet.delete_rows(row, 1)
    await message.channel.send('Old homework deleted!')


# Messageable.send has to be awaited while using discord
# function has to be asynchronous to use await
async def date_comparison(rows):
    """
    Compares the current date with the assignmnent due date, and checks whether the date has passed already
    *** I realized that this might be redundant with date_limitation ***
    :return:
    """
    # iteration starts at 1 with openpyxl, we are comparing the same indices(month)
    # if due-month > current month or due-day > current-day given the same month
    return int(sheet.cell(rows, 4).value[1:3]) > int(printable_date[0:2]) or int(sheet.cell(rows, 4).value[1:3]) \
           == int(printable_date[0:2] and int(sheet.cell(rows, 4).value[4:6]) > int(printable_date[3:5]))


async def date_limitation(rows):
    """
    checks to see if the assignment due dates fall within a certain, specified range

    :return:
    """
    # the number after the <= comparitor is within how many days you want to inform the user of due dates
    return (((int(sheet.cell(rows, 4).value[1:3]) - int(printable_date[0:2])) * 30) +
            (int(sheet.cell(rows, 4).value[4:6]) - int(printable_date[3:5]))) <= 10


async def convert_to_panda_tables(rows, message):
    """
    Converts excel sheets into panda tables

    :return:
    """
    # create dictionary of relevant values to use for dataframe
    course_details = ({
        "Name": [sheet.cell(rows, 1).value],
        "#": [sheet.cell(rows, 2).value],
        "Assignment": [sheet.cell(rows, 3).value],
        "Difficulty": [sheet.cell(rows, 5).value],
    })
    # initialize dataframe to use with pandas to create a table
    # all the keys are scalar values, so we must start with an index. Tried to use something non-intrusive
    dataframe = pd.DataFrame(course_details, columns=["Name", "#", "Assignment", "Difficulty"], index=["course"])
    # this is for discord formatting, as otherwise the spacing is off after conversion. Ticks don't
    # look the best, but I'm not sure what else to use
    await message.channel.send('\```' + dataframe.to_string() + '')
    # excel auto formats to non-numerical date without 1 space, so I'm stripping the space
    # could git rid of the strip() if someone knows how to properly format excel dates without auto-conversion
    await message.channel.send(f'Due: {sheet.cell(rows, 4).value.strip()}')


@client.event
async def on_message(message):
    """
    When a user enters a message, check to see if they are calling the bot using a key word, and if so, print out
    the relevant information

    :param message:
    :return:
    """
    # # what does this do? is it necessary?
    # if message.author == client.user:
    #     return
    if message.content.lower() == "deadlines" or message.content.lower() == "due".lower():
        # await date_comparison(message)
        for rows in range(2, sheet.max_row + 1):
            # just reading the date here
            for columns in range(4, 5):
                if await date_comparison(rows):
                    if await date_limitation(rows):
                        await convert_to_panda_tables(rows, message)
        await message.channel.send(random.choice(encouragements))

    # AttributeError: 'builtin_function_or_method' object has no attribute 'lower'
    # second part of conditional checks for a specific course
    elif message.content.startswith("deadlines") and len(message.content) == 14:
        for rows in range(2, sheet.max_row + 1):
            for columns in range(2, 3):
                cell_value = sheet.cell(row=rows, column=columns)
                if str(cell_value.value) == (message.content[10:]):
                    if await date_comparison(rows):
                        if await date_limitation(rows):
                            await convert_to_panda_tables(rows, message)

    elif message.content.lower() == 'cleanup':
        await delete_empty_rows(message)
        await delete_past_homework(rows, message)

load_dotenv()
TOKEN = os.getenv('DISCORD_TOKEN')
client.run(TOKEN)
